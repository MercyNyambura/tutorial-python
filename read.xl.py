"""import pandas as pd
df = pd.read_csv(r'c:\Users\Edgar\OneDrive\Documents\simple_csv.csv')

print(df)"""


"""import openpyxl
file = openpyxl.load_workbook("c:\Users\Edgar\Downloads\CovidDeaths.xlsx") 
print (file)"""




"""import openpyxl

invfile = openpyxl.load_workbook("c:\Users\Edgar\Downloads\CovidDeaths.xlsx")

product_list = inv_file("sheet1")

products_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name] = 1
        products_per_supplier[supplier_name] = current_num_products + 1

    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)"""

"""import xlrd
wb = xlrd.open_workbook("c:\Users\Edgar\Downloads\CovidDeaths.xlsx")
print(wb)"""




"""import xlrd

wb = xlrd.open_workbook('inventory.xlsx')
ws = wb.sheet_by_index(0)
for i in range(ws.nrows):
    for j in range(ws.ncols):
        print(ws.cell_value(i,j), end = '\t')
    print('')"""



import requests
"""import mysql
import pandas as pd
import warnings
warnings.filterwarnings('ignore')
connection = mysql.connect('travel.mysql')
cursor = connection.cursor()
cursor.execute('''select name from mysql where type = 'table';''')
print('list of tables present in the database')                               
table_list = [table[0] for table in cursor.fetchall()]
table_list"""             
                            
                                
                        








